"""
将修正后的CSV数据转换为Excel格式（含格式化）
"""

import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def format_percentage_fill(value):
    """根据百分比值返回背景颜色"""
    try:
        pct = float(value)
        if pct >= 85:
            return PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # 绿色
        elif pct >= 70:
            return PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 黄色
        elif pct >= 50:
            return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 浅红
        else:
            return PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 红色
    except:
        return PatternFill()

def create_excel_from_csv(csv_path, excel_path, sheet_name="统计"):
    """从CSV文件创建Excel"""
    
    # 读取CSV
    data = []
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for row in reader:
            data.append(row)
    
    # 创建Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入数据
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            
            # 数字格式化
            if row_idx > 1:  # 非标题行
                try:
                    # 尝试转换为浮点数
                    float_val = float(value)
                    # 根据列号判断小数位
                    if col_idx >= 5:  # 数字列
                        cell.number_format = '0.0' if '.' in value else '0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                except:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 标题行格式
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 合格率列格式化（最后一列）
            if '合格率' in data[0] and col_idx == len(data[0]):
                if row_idx > 1 and value:
                    cell.fill = format_percentage_fill(value)
                    cell.font = Font(bold=True)
    
    # 自动调整列宽
    for col_idx, row_data in enumerate(data[0], 1):
        max_width = len(str(row_data))
        for row_data_list in data[1:]:
            if col_idx <= len(row_data_list):
                max_width = max(max_width, len(str(row_data_list[col_idx - 1])))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 3, 20)
    
    # 保存
    wb.save(excel_path)
    print(f"✓ 已生成: {excel_path}")

def main():
    print("=" * 80)
    print("生成修正版Excel文件")
    print("=" * 80)
    
    # 生成按日期的Excel
    create_excel_from_csv(
        '13台设备统计_按日期_3月23日-4月10日_修正.csv',
        '13台设备统计_按日期_3月23日-4月10日_修正.xlsx',
        '按日期统计'
    )
    
    # 生成按设备的Excel
    create_excel_from_csv(
        '13台设备统计_修正版_3月23日-4月10日.csv',
        '13台设备统计_修正版_3月23日-4月10日.xlsx',
        '按设备统计'
    )
    
    print("\n✓ 所有Excel文件已生成完毕")

if __name__ == '__main__':
    main()
