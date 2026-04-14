import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("正在生成 Excel 文件...")

# 读取 CSV 文件
data_file = '8台设备问题数据_02-02到03-31.csv'
daily_file = '8台设备每日统计_02-02到03-31.csv'

df_data = pd.read_csv(data_file)
df_daily = pd.read_csv(daily_file)

# 创建 Excel writer
excel_file = '8台设备问题数据_02-02到03-31.xlsx'
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    # 写入详细数据（只保留关键列）
    key_columns = ['采集日期', '采集时间', '设备ID', '审批状态', '原始上送时长', 
                   '无效时长', '无效片段标记原因', '向下镜头视频链接', '向前镜头视频链接']
    df_data_export = df_data[[col for col in key_columns if col in df_data.columns]]
    df_data_export.to_excel(writer, sheet_name='详细数据', index=False)
    
    # 写入每日统计
    df_daily.to_excel(writer, sheet_name='每日统计', index=False)

# 应用格式化
workbook = load_workbook(excel_file)

# 格式化详细数据工作表
ws_detail = workbook['详细数据']
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 设置表头
for cell in ws_detail[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# 调整列宽
for col_num, col_title in enumerate(df_data_export.columns, 1):
    column_letter = get_column_letter(col_num)
    if col_title in ['向下镜头视频链接', '向前镜头视频链接']:
        ws_detail.column_dimensions[column_letter].width = 50
    elif col_title in ['无效片段标记原因']:
        ws_detail.column_dimensions[column_letter].width = 30
    else:
        ws_detail.column_dimensions[column_letter].width = 15

# 赋予所有数据单元格边框
for row in ws_detail.iter_rows(min_row=2, max_row=len(df_data_export)+1, min_col=1, max_col=len(df_data_export.columns)):
    for cell in row:
        cell.border = border
        if cell.column in [5, 6]:  # 时长列
            cell.number_format = '0.00'

# 格式化每日统计工作表
ws_daily = workbook['每日统计']
for cell in ws_daily[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

for col_num in range(1, len(df_daily.columns)+1):
    column_letter = get_column_letter(col_num)
    ws_daily.column_dimensions[column_letter].width = 15

# 赋予所有数据单元格边框
for row in ws_daily.iter_rows(min_row=2, max_row=len(df_daily)+1, min_col=1, max_col=len(df_daily.columns)):
    for cell in row:
        cell.border = border
        if cell.column in [4, 5]:  # 时长列
            cell.number_format = '0.00'

workbook.save(excel_file)
print(f"✓ 已生成: {excel_file}")

print(f"\n=== 统计摘要 ===")
print(f"总数据行数: {len(df_data)}")
print(f"涉及天数: {len(df_daily)}")
print(f"涉及设备数: {df_data['设备ID'].nunique()}")
print(f"日期范围: {df_daily['采集日期'].min()} ~ {df_daily['采集日期'].max()}")
print(f"总无效时长(分): {df_daily['无效时长(分)'].sum():.2f}")
